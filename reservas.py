#!/usr/bin/env python3
from datetime import datetime, date, timedelta
import sqlite3
from typing import Optional, List, Tuple
from tabulate import tabulate
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

DB_FILE = "reservas.db"
TURNOS = ["Matutino", "Vespertino", "Nocturno"]
DATE_FORMAT = "%m-%d-%Y"

def obtener_conexion():
    """Retorna una conexión a la base de datos SQLite."""
    return sqlite3.connect(DB_FILE)

def crear_tablas():
    """Crea las tablas necesarias si no existen."""
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS clientes(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        apellidos TEXT NOT NULL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS salas(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        cupo INTEGER NOT NULL CHECK(cupo>0)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS reservaciones(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        folio TEXT UNIQUE,
        cliente_id INTEGER NOT NULL,
        sala_id INTEGER NOT NULL,
        fecha TEXT NOT NULL,
        turno TEXT NOT NULL,
        nombre_evento TEXT NOT NULL,
        FOREIGN KEY(cliente_id) REFERENCES clientes(id),
        FOREIGN KEY(sala_id) REFERENCES salas(id),
        UNIQUE(sala_id, fecha, turno)
    )""")
    conn.commit()
    conn.close()

def iniciar_estado():
    """Informa si existe estado previo o si se inicia vacío."""
    try:
        conn = obtener_conexion()
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clientes'")
        existe = cur.fetchone() is not None
        conn.close()
        if not existe:
            print("Iniciando con estado inicial vacío.")
        else:
            print("Estado previo encontrado. Cargando datos...")
    except Exception as e:
        print("Iniciando con estado inicial vacío.")

def listar_clientes() -> List[Tuple[int,str,str]]:
    """Retorna lista de clientes ordenada por apellidos, nombre."""
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, apellidos FROM clientes ORDER BY apellidos, nombre")
    rows = cur.fetchall()
    conn.close()
    return rows

def mostrar_clientes():
    """Muestra la tabla de clientes con sus claves."""
    rows = listar_clientes()
    if not rows:
        print("No hay clientes registrados.")
        return
    table = [(r[0], r[2], r[1]) for r in rows]
    print(tabulate(table, headers=["Clave","Apellidos","Nombre"], tablefmt="github"))

def registrar_cliente():
    """Registra un nuevo cliente pidiendo nombre y apellidos."""
    while True:
        nombre = input("Nombre del cliente: ").strip()
        if nombre:
            break
        print("El nombre no puede quedar vacío.")
    while True:
        apellidos = input("Apellidos del cliente: ").strip()
        if apellidos:
            break
        print("Los apellidos no pueden quedar vacíos.")
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("INSERT INTO clientes(nombre, apellidos) VALUES(?,?)", (nombre, apellidos))
    conn.commit()
    clave = cur.lastrowid
    conn.close()
    print(f"Cliente registrado con clave: {clave}")

def listar_salas() -> List[Tuple[int,str,int]]:
    """Retorna lista de salas."""
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, cupo FROM salas ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    return rows

def mostrar_salas():
    """Muestra las salas con clave y cupo."""
    rows = listar_salas()
    if not rows:
        print("No hay salas registradas.")
        return
    print(tabulate(rows, headers=["Clave","Nombre","Cupo"], tablefmt="github"))

def registrar_sala():
    """Registra una sala nueva pidiendo nombre y cupo."""
    while True:
        nombre = input("Nombre de la sala: ").strip()
        if nombre:
            break
        print("El nombre no puede quedar vacío.")
    while True:
        try:
            cupo = int(input("Cupo de la sala (entero > 0): ").strip())
            if cupo > 0:
                break
            else:
                print("El cupo debe ser mayor a cero.")
        except:
            print("Ingrese un número entero válido.")
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("INSERT INTO salas(nombre, cupo) VALUES(?,?)", (nombre, cupo))
    conn.commit()
    clave = cur.lastrowid
    conn.close()
    print(f"Sala registrada con clave: {clave}")

def parsear_fecha(texto: str) -> Optional[date]:
    """Convierte texto mm-dd-aaaa a objeto date o retorna None."""
    try:
        dt = datetime.strptime(texto, DATE_FORMAT).date()
        return dt
    except:
        return None

def solicitar_fecha_minima() -> date:
    """Retorna la fecha mínima permitida (hoy + 2 días)."""
    return date.today() + timedelta(days=2)

def es_domingo(d: date) -> bool:
    """Determina si la fecha es domingo."""
    return d.weekday() == 6

def siguiente_lunes(d: date) -> date:
    """Dado domingo devuelve lunes siguiente, sino retorna mismo día."""
    if is_sunday := d.weekday() == 6:
        return d + timedelta(days=1)
    return d

def salas_disponibles_para_fecha(d: date) -> List[Tuple[int,str,int,List[str]]]:
    """Retorna salas que tengan al menos un turno disponible para la fecha dada y qué turnos."""
    conn = obtener_conexion()
    cur = conn.cursor()
    salas = listar_salas()
    disponibles = []
    for s in salas:
        s_id, s_nombre, s_cupo = s
        turnos_libres = []
        for t in TURNOS:
            cur.execute("SELECT COUNT(1) FROM reservaciones WHERE sala_id=? AND fecha=? AND turno=?", (s_id, d.strftime(DATE_FORMAT), t))
            cnt = cur.fetchone()[0]
            if cnt == 0:
                turnos_libres.append(t)
        if turnos_libres:
            disponibles.append((s_id, s_nombre, s_cupo, turnos_libres))
    conn.close()
    return disponibles

def generar_folio(res_id: int) -> str:
    """Genera folio a partir del id de la reservación."""
    return f"R{res_id:06d}"

def registrar_reservacion():
    """Flujo para registrar una reservación con todas las validaciones."""
    clientes = listar_clientes()
    if not clientes:
        print("No hay clientes registrados. Registra uno primero.")
        return
    while True:
        mostrar_clientes()
        try:
            clave = input("Selecciona la clave del cliente (o 'C' para cancelar): ").strip()
            if clave.lower() == 'c':
                print("Operación cancelada.")
                return
            cliente_id = int(clave)
            if any(c[0] == cliente_id for c in clientes):
                break
            else:
                print("Clave no existe. Intenta de nuevo.")
        except:
            print("Entrada inválida.")
    min_fecha = solicitar_fecha_minima()
    while True:
        entrada = input(f"Ingrese la fecha a reservar en formato mm-dd-aaaa (mínimo {min_fecha.strftime(DATE_FORMAT)}): ").strip()
        d = parsear_fecha(entrada)
        if not d:
            print("Formato de fecha inválido.")
            continue
        if d < min_fecha:
            print("La fecha debe ser por lo menos dos días después de la fecha actual.")
            continue
        if es_domingo(d):
            propuesta = d + timedelta(days=1)
            print(f"La fecha indicada es domingo. Se propone automáticamente el lunes {propuesta.strftime(DATE_FORMAT)}.")
            aceptar = input(f"¿Aceptar {propuesta.strftime(DATE_FORMAT)}? (S/N): ").strip().lower()
            if aceptar == 's':
                d = propuesta
            else:
                continue
        disponibles = salas_disponibles_para_fecha(d)
        if not disponibles:
            print("No hay salas disponibles para esa fecha. Elige otra fecha.")
            continue
        break
    print("Salas con turnos disponibles para la fecha:")
    table = []
    for s in disponibles:
        table.append((s[0], s[1], s[2], ", ".join(s[3])))
    print(tabulate(table, headers=["Clave","Sala","Cupo","Turnos disponibles"], tablefmt="github"))
    while True:
        try:
            sala_elegida = int(input("Selecciona la clave de la sala (o '0' para cancelar): ").strip())
            if sala_elegida == 0:
                print("Operación cancelada.")
                return
            if any(s[0] == sala_elegida for s in disponibles):
                break
            else:
                print("Clave de sala inválida.")
        except:
            print("Entrada inválida.")
    sala_turnos = next(s[3] for s in disponibles if s[0] == sala_elegida)
    print("Turnos disponibles para la sala seleccionada:", ", ".join(sala_turnos))
    while True:
        turno = input(f"Seleccione turno ({'/'.join(sala_turnos)}): ").strip()
        if turno in sala_turnos:
            break
        print("Turno inválido.")
    while True:
        nombre_evento = input("Nombre del evento: ").strip()
        if nombre_evento and len(nombre_evento.strip())>0:
            break
        print("El nombre del evento no puede quedar vacío.")
    conn = obtener_conexion()
    cur = conn.cursor()
    fecha_str = d.strftime(DATE_FORMAT)
    try:
        cur.execute("INSERT INTO reservaciones(folio, cliente_id, sala_id, fecha, turno, nombre_evento) VALUES(?,?,?,?,?,?)",
                    ("", cliente_id, sala_elegida, fecha_str, turno, nombre_evento))
        conn.commit()
        res_id = cur.lastrowid
        folio = generar_folio(res_id)
        cur.execute("UPDATE reservaciones SET folio=? WHERE id=?", (folio, res_id))
        conn.commit()
        print(f"Reservación registrada. Folio: {folio}")
    except sqlite3.IntegrityError:
        print("Error: La sala ya está reservada para ese turno y fecha.")
    finally:
        conn.close()

def editar_nombre_evento():
    """Edita el nombre del evento de una reservación dentro de un rango de fechas."""
    print("Proporcione rango de fechas para buscar eventos.")
    while True:
        desde_txt = input("Fecha desde (mm-dd-aaaa): ").strip()
        desde = parsear_fecha(desde_txt)
        if desde:
            break
        print("Formato inválido.")
    while True:
        hasta_txt = input("Fecha hasta (mm-dd-aaaa): ").strip()
        hasta = parsear_fecha(hasta_txt)
        if hasta and hasta >= desde:
            break
        print("Formato inválido o fecha final anterior a inicial.")
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("SELECT id, folio, nombre_evento, fecha FROM reservaciones WHERE fecha BETWEEN ? AND ? ORDER BY fecha",
                (desde.strftime(DATE_FORMAT), hasta.strftime(DATE_FORMAT)))
    rows = cur.fetchall()
    if not rows:
        print("No hay eventos en ese rango.")
        conn.close()
        return
    table = [(r[0], r[1], r[2], r[3]) for r in rows]
    print(tabulate(table, headers=["ID","Folio","Nombre Evento","Fecha"], tablefmt="github"))
    while True:
        clave = input("Indique la clave (ID) del evento a modificar o 'C' para cancelar: ").strip()
        if clave.lower() == 'c':
            print("Operación cancelada.")
            conn.close()
            return
        try:
            clave_id = int(clave)
            if any(r[0] == clave_id for r in rows):
                break
            else:
                print("Clave no pertenece al listado mostrado.")
        except:
            print("Entrada inválida.")
    while True:
        nuevo_nombre = input("Nuevo nombre del evento: ").strip()
        if nuevo_nombre:
            break
        print("El nombre del evento no puede quedar vacío.")
    cur.execute("UPDATE reservaciones SET nombre_evento=? WHERE id=?", (nuevo_nombre, clave_id))
    conn.commit()
    conn.close()
    print("Nombre del evento actualizado.")

def consultar_reservaciones():
    """Consulta reservaciones para una fecha específica (o hoy si se omite)."""
    entrada = input("Ingrese fecha mm-dd-aaaa para consultar (ENTER = fecha actual): ").strip()
    if not entrada:
        d = date.today()
    else:
        d = parsear_fecha(entrada)
        if not d:
            print("Formato de fecha inválido. Usando fecha actual.")
            d = date.today()
    fecha_str = d.strftime(DATE_FORMAT)
    conn = obtener_conexion()
    cur = conn.cursor()
    cur.execute("""SELECT r.folio, r.nombre_evento, r.fecha, r.turno, s.nombre, s.cupo, c.apellidos || ', ' || c.nombre
                   FROM reservaciones r
                   JOIN salas s ON r.sala_id = s.id
                   JOIN clientes c ON r.cliente_id = c.id
                   WHERE r.fecha = ?
                   ORDER BY r.turno, s.id""", (fecha_str,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        print("No hay reservaciones para esa fecha.")
        return
    table = []
    for r in rows:
        table.append((r[0], r[1], r[2], r[3], r[4], r[5], r[6]))
    print(tabulate(table, headers=["Folio","Evento","Fecha","Turno","Sala","Cupo","Cliente"], tablefmt="github"))
    while True:
        opc = input("¿Desea exportar este reporte? (S/N): ").strip().lower()
        if opc == 's':
            exportar_reporte(table)
            break
        elif opc == 'n':
            break
        else:
            print("Opción inválida.")

def exportar_reporte(tabla: List[Tuple]):
    """Exporta el listado a CSV, JSON o Excel con el formato solicitado."""
    df = pd.DataFrame(tabla, columns=["Folio","Evento","Fecha","Turno","Sala","Cupo","Cliente"])
    while True:
        opc = input("Formato de exportación: 1) CSV  2) JSON  3) Excel  (Elige 1/2/3 o C para cancelar): ").strip().lower()
        if opc == 'c':
            return
        if opc == '1':
            nombre = input("Nombre archivo CSV (ej: reporte.csv): ").strip() or "reporte.csv"
            df.to_csv(nombre, index=False)
            print(f"Exportado a {nombre}")
            return
        if opc == '2':
            nombre = input("Nombre archivo JSON (ej: reporte.json): ").strip() or "reporte.json"
            df.to_json(nombre, orient="records", date_format="iso", force_ascii=False)
            print(f"Exportado a {nombre}")
            return
        if opc == '3':
            nombre = input("Nombre archivo Excel (ej: reporte.xlsx): ").strip() or "reporte.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            headers = list(df.columns)
            ws.append(headers)
            for row in df.itertuples(index=False):
                ws.append(list(row))
            bold = Font(bold=True)
            thick = Side(border_style="thick")
            border = Border(bottom=thick)
            for col_idx in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold
                cell.border = border
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(headers), max_row=1+len(df)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            wb.save(nombre)
            print(f"Exportado a {nombre} con formato requerido.")
            return
        print("Opción inválida.")

def confirmar_salida() -> bool:
    """Solicita confirmación antes de salir."""
    opc = input("¿Confirmar salida? (S/N): ").strip().lower()
    return opc == 's'

def menu_principal():
    """Muestra el menú principal y dirige a las funciones correspondientes."""
    while True:
        print("\n--- Menú Principal ---")
        print("1) Registrar reservación de sala")
        print("2) Editar nombre de evento")
        print("3) Consultar reservaciones por fecha")
        print("4) Registrar nuevo cliente")
        print("5) Registrar nueva sala")
        print("6) Listar clientes")
        print("7) Listar salas")
        print("8) Salir")
        opcion = input("Selecciona una opción (1-8): ").strip()
        if opcion == '1':
            registrar_reservacion()
        elif opcion == '2':
            editar_nombre_evento()
        elif opcion == '3':
            consultar_reservaciones()
        elif opcion == '4':
            registrar_cliente()
        elif opcion == '5':
            registrar_sala()
        elif opcion == '6':
            mostrar_clientes()
        elif opcion == '7':
            mostrar_salas()
        elif opcion == '8':
            if confirmar_salida():
                print("Saliendo del sistema.")
                break
            else:
                print("Regresando al menú principal.")
        else:
            print("Opción inválida.")

def main():
    """Punto de entrada del programa."""
    crear_tablas()
    iniciar_estado()
    menu_principal()

if __name__ == "__main__":
    main()
